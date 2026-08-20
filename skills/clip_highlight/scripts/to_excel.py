# clip_highlight/scripts/to_excel.py
"""把分析记录转成结构化表格：analysis.md -> JSON -> xlsx

JSON 解析失败时把 json 模块的报错原文回送给模型让它重出 ——
报错里带着出错位置(行号/列号/字符偏移)，比笼统说一句「格式不对」有用得多。
"""
import argparse
import json
import re
import sys
import traceback
from datetime import date
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from clear_tmp import clear_tmp
from srt_io import configure_utf8_stdio, read_text

DEFAULT_MAX_REPAIR = 2
DEFAULT_OUTPUT_DIR = Path("outputs")

# 内置文件跟着脚本走，不能用 CWD 相对路径
SKILL_DIR = Path(__file__).resolve().parent.parent
ASSETS_DIR = SKILL_DIR / "assets"
EXCEL_PROMPT = ASSETS_DIR / "to_excel" / "excel_prompt.md"

# 模型仍可能把 JSON 包在代码块里。这是确定性能修的，
# 没必要为它多烧一次 API 调用。
FENCE = re.compile(r"```(?:json)?\s*(.*?)\s*```", re.DOTALL)

COLUMNS = [
    ("start", "开始时间", 12),
    ("end", "结束时间", 12),
    ("category", "话题分类", 18),
    ("detail", "内容详情摘要", 80),
    ("highlight", "对应切片标题/高能点", 60),
    ("editor", "剪辑", 10),
]


def parse_args():
    """配置命令行参数解析"""
    parser = argparse.ArgumentParser(
        prog="to_excel.py",
        description="Turn an analysis markdown file into JSON and an xlsx sheet.",
    )
    parser.add_argument(
        "analysis",
        type=Path,
        nargs="?",
        help="Path to the analysis markdown file (omit it when using --from-json)",
    )
    parser.add_argument(
        "--from-json",
        type=Path,
        default=None,
        help="Rebuild the xlsx from an existing rows.json, skipping the LLM entirely. "
             "Use this after hand-fixing a row -- re-running the analysis would "
             "re-generate every other row too.",
    )
    parser.add_argument(
        "--sheet-date",
        default="",
        help="Stream date as 8 digits (YYYYMMDD). Defaults to today.",
    )
    parser.add_argument(
        "--max-repair",
        type=int,
        default=DEFAULT_MAX_REPAIR,
        help=f"Rounds to send a JSON error back for a retry (default: {DEFAULT_MAX_REPAIR})",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help=f"Where to write the results (default: {DEFAULT_OUTPUT_DIR.as_posix()})",
    )
    return parser.parse_args()


def extract_json(response: str) -> str:
    """剥掉可能存在的 ``` 代码块围栏，留下裸 JSON"""
    match = FENCE.search(response)
    return match.group(1) if match else response.strip()


def request_json(analysis_text: str, system_prompt: str, max_repair: int) -> Optional[dict]:
    """要一份 JSON，解析失败就把报错回送让模型自己修"""
    # 延后导入：llm.py 一加载就会初始化 genai client(要 API key)，
    # 而 --from-json 只是拿现成的 JSON 重出表格，不该被凭证卡住。
    from llm import MODEL, call_llm, client

    user_prompt = f"【直播全场记录】\n{analysis_text}"

    for attempt in range(max_repair + 1):
        response, _ = call_llm(client, MODEL, system_prompt, user_prompt)
        candidate = extract_json(response)

        try:
            return json.loads(candidate)
        except json.JSONDecodeError as error:
            if attempt == max_repair:
                print(f"[Error] JSON still invalid after {max_repair} repair round(s): {error}")
                return None

            print(f"[Warning] Invalid JSON ({error}), sending the error back for a retry")
            # 把出错的原文一并回送 —— 只说「错了」模型无从下手，
            # 给它自己的输出加上报错位置才改得动。
            user_prompt = (
                f"你上一次的输出不是合法 JSON，Python 的 json 模块报错如下：\n"
                f"{error}\n\n"
                f"以下是你上一次的输出，请修正该错误后**只输出修正后的完整 JSON**，"
                f"不要输出任何解释：\n{candidate}"
            )

    return None


def normalise(payload: dict, sheet_date: str) -> List[Dict[str, Any]]:
    """校验并整理成行列表，顺便把该报的问题报出来"""
    rows = payload.get("rows")
    if not isinstance(rows, list) or not rows:
        print("[Error] Response has no usable 'rows' array")
        return []

    cleaned = []
    for position, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            print(f"[Warning] Row #{position} is not an object, skipping")
            continue
        # 缺键补空字符串而不是丢行 —— 丢行是信息损失，空格子人工一眼就能看见
        missing = [key for key, _, _ in COLUMNS if key not in row]
        if missing:
            print(f"[Warning] Row #{position} missing field(s): {', '.join(missing)}")
        cleaned.append({key: str(row.get(key, "") or "") for key, _, _ in COLUMNS})

    print(f"[Info] {len(cleaned)} row(s) parsed, sheet_date={sheet_date}")
    return cleaned


def write_xlsx(rows: List[Dict[str, Any]], sheet_date: str, output_path: Path) -> Path:
    """把行列表写成 xlsx

    <br> 是给表格单元格用的换行标记，写进 Excel 时还原成真实换行，
    并给单元格开自动换行，否则长详情会糊成一行看不了。
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_date or "sheet"

    header_font = Font(bold=True)
    for column_index, (_, title, width) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=1, column=column_index, value=title)
        cell.font = header_font
        sheet.column_dimensions[get_column_letter(column_index)].width = width
    sheet.freeze_panes = "A2"

    for row_index, row in enumerate(rows, start=2):
        for column_index, (key, _, _) in enumerate(COLUMNS, start=1):
            cell = sheet.cell(
                row=row_index, column=column_index, value=row[key].replace("<br>", "\n")
            )
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def main() -> int:
    configure_utf8_stdio()
    args = parse_args()

    if args.from_json:
        # 修表路径：JSON 已经在手上，跳过模型。重跑分析会把没问题的行
        # 也一起重出，改一格却动了整张表，反而更难对。
        raw = read_text(args.from_json)
        if not raw:
            return 1
        try:
            payload = json.loads(raw)
        except json.JSONDecodeError as error:
            print(f"[Error] {args.from_json} is not valid JSON: {error}")
            return 1
        source = args.from_json
    else:
        if args.analysis is None:
            print("[Error] Pass an analysis markdown file, or --from-json <rows.json>")
            return 2

        analysis_text = read_text(args.analysis)
        if not analysis_text:
            return 1

        system_prompt = read_text(EXCEL_PROMPT)
        if not system_prompt:
            return 1

        payload = request_json(analysis_text, system_prompt, args.max_repair)
        if payload is None:
            return 1
        source = args.analysis

    # 文件名形如 test.analysis.md / test.rows.json，
    # 取最前面的 stem 当这场直播的标识
    stem = source.name.split(".")[0]
    sheet_date = args.sheet_date or str(payload.get("sheet_date") or "").strip()
    if not re.fullmatch(r"\d{8}", sheet_date):
        sheet_date = date.today().strftime("%Y%m%d")

    rows = normalise(payload, sheet_date)
    if not rows:
        return 1

    stream_dir = args.output_dir / stem
    stream_dir.mkdir(parents=True, exist_ok=True)

    json_path = stream_dir / f"{stem}.rows.json"
    json_path.write_text(
        json.dumps({"sheet_date": sheet_date, "rows": rows}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"[Done] JSON written to {json_path}")

    xlsx_path = write_xlsx(rows, sheet_date, stream_dir / f"{stem}.xlsx")
    print(f"[Done] Spreadsheet written to {xlsx_path}")

    # 流水线最后一步，中间产物到这里就没用了。
    # 交付物都在 outputs/ 下，不受影响。
    # --from-json 是事后修表，不是跑流水线 —— 别顺手清掉人家还在看的 tmp/。
    # 清不掉不影响交付 —— xlsx 上一行已经写完了，报一句就收工。
    if not args.from_json and clear_tmp() < 0:
        print("[Warning] tmp/ 没清干净，但 xlsx 已经写好了，不影响这次交付。")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception:
        print("\n" + "!" * 60, file=sys.stderr)
        print("[Fatal] The program crashed while running:", file=sys.stderr)
        traceback.print_exc()
        print("!" * 60, file=sys.stderr)
        sys.exit(1)
